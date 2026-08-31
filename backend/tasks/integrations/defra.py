"""
DEFRA / DESNZ UK Government GHG Conversion Factors importer.

This is the library that makes the product usable without a Climatiq key. The
Climatiq integration next door can only *refresh* rows that already carry a
ClimatiqActivityId, so it cannot populate an empty database; nothing else ever
seeded EmissionFactors, which is why production shipped with zero factors and
users could not create emissions at all.

Source
------
GOV.UK publishes the conversion factors annually under the Open Government
Licence v3.0 (free reuse, including commercially, with attribution). Alongside
the human-readable workbook it publishes a **flat file "for automatic processing
only"** — a single normalised sheet, which is what this module consumes. No
scraping is involved: the attachment list is available as JSON from the GOV.UK
content API, so the download URL is resolved at runtime rather than hard-coded.
That matters because the asset URL carries a content hash and changes whenever
the file is revised (the 2026 edition was revised in July 2026).

Resolution order for the source, most explicit first:
    1. an explicit local path (offline / pinned imports, and the test suite)
    2. an explicit URL
    3. settings.DEFRA_EF_SPREADSHEET_URL
    4. the GOV.UK content API for the requested year

Deliberately NOT reusing ClimatiqActivityId
-------------------------------------------
It would be tempting to store the DEFRA row ID there for idempotency. Doing so
would be a live corruption bug: sync_climatiq_emission_factors selects rows with
`.exclude(ClimatiqActivityId="")` and overwrites their FactorValue from the
Climatiq API. DEFRA rows must keep that field empty so the Climatiq sync cannot
see them.
"""
import logging
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import transaction
from django.utils.timezone import now

import requests

from celery import shared_task

logger = logging.getLogger(__name__)

TIMEOUT = 120
USER_AGENT = "SusDevOS emission-factor importer"

GOVUK_CONTENT_API = (
    "https://www.gov.uk/api/content/government/publications/"
    "greenhouse-gas-reporting-conversion-factors-{year}"
)

# The edition imported when no year is given. Bump when DEFRA publishes.
DEFAULT_YEAR = 2026

DATA_SHEET = "Factors by Category"
PUBLISHER = "DEFRA / DESNZ"
LICENCE = "Open Government Licence v3.0"

# Only the aggregate rows become factors. The file also carries a per-gas
# breakdown of each activity ("kg CO2e of CH4 per unit" and friends); importing
# those as separate rows would multiply the picker four-fold and invite users to
# select a component instead of the total.
AGGREGATE_GHG_UNIT = "kg CO2e"

# The columns of the flat file, in order, as published. The value column is
# year-suffixed ("GHG Conversion Factor 2026"), so it is matched by prefix while
# the rest must match exactly.
EXPECTED_HEADER = [
    "ID", "Scope", "Level 1", "Level 2", "Level 3", "Level 4",
    "Column Text", "UOM", "GHG/Unit",
]
VALUE_COLUMN_PREFIX = "GHG Conversion Factor"
COLUMN_COUNT = len(EXPECTED_HEADER) + 1


class DefraImportError(RuntimeError):
    """Raised when the source cannot be fetched or does not look like the flat file."""


# ── Source resolution ────────────────────────────────────────────────────────


def resolve_flat_file_url(year: int) -> str:
    """Find the flat-file attachment URL for a given year via the GOV.UK content API."""
    url = GOVUK_CONTENT_API.format(year=year)
    try:
        resp = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
        resp.raise_for_status()
        payload = resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise DefraImportError(f"Could not read the GOV.UK content API for {year}: {exc}") from exc

    attachments = (payload.get("details") or {}).get("attachments") or []
    for attachment in attachments:
        title = (attachment.get("title") or "").lower()
        href = attachment.get("url") or ""
        if "flat file" in title and href.endswith((".xlsx", ".xls")):
            return href

    titles = [a.get("title") for a in attachments]
    raise DefraImportError(
        f"No 'flat file' spreadsheet attachment published for {year}. Attachments: {titles}"
    )


def download(url: str) -> bytes:
    try:
        resp = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise DefraImportError(f"Could not download {url}: {exc}") from exc
    return resp.content


# ── Parsing ──────────────────────────────────────────────────────────────────


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _scope_number(raw: str):
    """'Scope 1' -> 1. Returns None for anything else ('Outside of scopes')."""
    digits = "".join(c for c in raw if c.isdigit())
    if not digits:
        return None
    scope = int(digits)
    return scope if scope in (1, 2, 3) else None


def _activity_name(level2: str, level3: str, level4: str, column_text: str) -> str:
    """The specific label, most general part first, empties dropped.

    Level 1 is held separately as ActivityCategory, so it is not repeated here.
    """
    parts = [p for p in (level2, level3, level4, column_text) if p]
    return " - ".join(parts)[:200]


def parse_flat_file(content: bytes) -> list[dict]:
    """Parse the flat file into aggregate CO2e factor rows.

    Raises DefraImportError if the sheet or header is not what we expect, rather
    than silently importing nothing or importing the wrong columns — a factor
    library that quietly imports garbage is worse than one that fails loudly.
    """
    import io

    import openpyxl

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of types on bad input
        raise DefraImportError(f"Could not open the workbook: {exc}") from exc

    if DATA_SHEET not in workbook.sheetnames:
        raise DefraImportError(
            f"Sheet {DATA_SHEET!r} not found. Sheets present: {workbook.sheetnames}"
        )

    sheet = workbook[DATA_SHEET]
    rows = sheet.iter_rows(values_only=True)

    header_index = None
    for index, row in enumerate(rows):
        if _clean(row[0] if row else "") == "ID":
            header = [_clean(c) for c in row[:COLUMN_COUNT]]
            if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
                raise DefraImportError(
                    "The flat file's columns have changed. "
                    f"Expected {EXPECTED_HEADER}, found {header[: len(EXPECTED_HEADER)]}."
                )
            value_column = header[len(EXPECTED_HEADER)] if len(header) > len(EXPECTED_HEADER) else ""
            if not value_column.startswith(VALUE_COLUMN_PREFIX):
                raise DefraImportError(
                    f"Expected a {VALUE_COLUMN_PREFIX!r} column, found {value_column!r}."
                )
            header_index = index
            break

    if header_index is None:
        raise DefraImportError("No header row starting with 'ID' found in the data sheet.")

    parsed, skipped = [], 0
    for row in rows:  # the iterator is already positioned after the header
        if not row or not _clean(row[0]):
            continue
        if _clean(row[8]) != AGGREGATE_GHG_UNIT:
            continue

        scope = _scope_number(_clean(row[1]))
        if scope is None:
            skipped += 1  # 'Outside of scopes' rows have no GHG Protocol scope
            continue

        try:
            value = Decimal(_clean(row[9]))
        except (InvalidOperation, ValueError):
            skipped += 1
            continue

        name = _activity_name(_clean(row[3]), _clean(row[4]), _clean(row[5]), _clean(row[6]))
        if not name:
            skipped += 1
            continue

        parsed.append({
            "defra_id": _clean(row[0]),
            "scope": scope,
            "category": _clean(row[2])[:200],
            "name": name,
            "uom": _clean(row[7]),
            "value": value,
        })

    logger.info("DEFRA flat file parsed: %d aggregate factors, %d rows skipped", len(parsed), skipped)
    return parsed


# ── Import ───────────────────────────────────────────────────────────────────


def import_factors(*, rows: list[dict], year: int, version: str = "") -> dict:
    """Upsert parsed rows into EmissionFactors under a DEFRA <year> set.

    Idempotent: keyed on (SetId, ActivityName, InputUnitId, Gas), so re-running
    updates values in place rather than duplicating the library.
    """
    from apps.emissions.models import EmissionFactors, EmissionFactorSets, Units

    if not rows:
        raise DefraImportError("Refusing to import an empty factor set.")

    units = {u.UnitName: u for u in Units.objects.all()}
    missing_units = sorted({r["uom"] for r in rows if r["uom"] and r["uom"] not in units})
    if missing_units:
        raise DefraImportError(
            "These units are in the DEFRA file but not in the Units table: "
            f"{missing_units}. Run `manage.py seed_units` first."
        )

    ef_set, _ = EmissionFactorSets.objects.update_or_create(
        SetName=f"DEFRA {year}",
        Publisher=PUBLISHER,
        defaults={
            "Version": version,
            "ApplicableYear": year,
            "GeographicScope": "United Kingdom",
            "IsActive": True,
        },
    )

    created = updated = 0
    timestamp = now()

    with transaction.atomic():
        for row in rows:
            # ActivityCategory (DEFRA "Level 1") is part of the key, not just a
            # payload field. Without it "Fuels / Gaseous fuels / Butane" and
            # "WTT- fuels / Gaseous fuels / Butane" collide - direct combustion
            # and well-to-tank share every other label - and one silently
            # overwrites the other. That collapsed 671 of 3425 rows.
            _, was_created = EmissionFactors.objects.update_or_create(
                SetId=ef_set,
                ActivityCategory=row["category"],
                ActivityName=row["name"],
                InputUnitId=units.get(row["uom"]),
                Gas="CO2e",
                defaults={
                    "Scope": row["scope"],
                    "FactorValue": row["value"],
                    # Provenance of the dataset (UK Government factors for UK
                    # reporters), not the geography of the activity itself.
                    "CountryCode": "GBR",
                    "ApplicableYear": year,
                    # Left empty on purpose - see the module docstring.
                    "ClimatiqActivityId": "",
                    "ExternalSyncedAt": timestamp,
                },
            )
            created += was_created
            updated += not was_created

        # Inside the atomic block on purpose, so raising actually rolls back.
        # A parsed row that does not become a stored row is a key collision -
        # two distinct published factors folded into one - and a silently
        # collapsed library hands users the wrong number with no visible symptom.
        stored = EmissionFactors.objects.filter(SetId=ef_set).count()
        if stored != len(rows):
            raise DefraImportError(
                f"Parsed {len(rows)} factors but {ef_set.SetName} would hold {stored}. "
                "Two published factors share an upsert key; the import was rolled back."
            )

    return {"set": ef_set.SetName, "created": created, "updated": updated, "total": len(rows)}


def run_import(*, year: int, url: str = "", path: str = "") -> dict:
    """Resolve the source, parse it, and import. Returns a summary dict."""
    if path:
        with open(path, "rb") as handle:
            content = handle.read()
        source = path
    else:
        source = url or getattr(settings, "DEFRA_EF_SPREADSHEET_URL", "") or resolve_flat_file_url(year)
        content = download(source)

    rows = parse_flat_file(content)
    result = import_factors(rows=rows, year=year)
    result["source"] = source
    result["licence"] = LICENCE
    return result


# ── Scheduled import ─────────────────────────────────────────────────────────


@shared_task(
    name="tasks.integrations.sync_defra_emission_factors",
    bind=True,
    max_retries=3,
    default_retry_delay=3600,
)
def sync_defra_emission_factors(self, year: int = 0):
    """Import or refresh the DEFRA factor library. Idempotent.

    Scheduled annually rather than weekly: DEFRA publishes each June and revises
    within a month or two, so this runs in mid-July. Unlike the Climatiq sync it
    can populate an empty database, which is the whole point - it is the path
    that works with no API key of any kind.

    Needs no credentials: the data is Open Government Licence v3.0.
    """
    target = year or DEFAULT_YEAR
    try:
        result = run_import(year=target)
    except DefraImportError as exc:
        logger.error("sync_defra_emission_factors: %s", exc)
        raise self.retry(exc=exc) from exc

    logger.info(
        "sync_defra_emission_factors: %s - %d created, %d updated",
        result["set"], result["created"], result["updated"],
    )
    return result
